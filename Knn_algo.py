from math import sqrt
from sklearn import neighbors
from os import listdir
from os.path import isdir, join, isfile, splitext
import pickle
from PIL import Image, ImageFont, ImageDraw, ImageEnhance
import face_recognition
from face_recognition import face_locations
from face_recognition.cli import image_files_in_folder
import cv2
from openpyxl import Workbook, load_workbook
import datetime
from tkinter import *
from tkinter.filedialog import  *
import tkinter.messagebox

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

def predict(X_img_path, knn_clf = None, model_save_path ="", DIST_THRESH = .45):
    """
    recognizes faces in given image, based on a trained knn classifier

    :param X_img_path: path to image to be recognized
    :param knn_clf: (optional) a knn classifier object. if not specified, model_save_path must be specified.
    :param model_save_path: (optional) path to a pickled knn classifier. if not specified, model_save_path must be knn_clf.
    :param DIST_THRESH: (optional) distance threshold in knn classification. the larger it is, the more chance of misclassifying an unknown person to a known one.
    :return: a list of names and face locations for the recognized faces in the image: [(name, bounding box), ...].
        For faces of unrecognized persons, the name 'N/A' will be passed.
    """

    if not isfile(X_img_path) or splitext(X_img_path)[1][1:] not in ALLOWED_EXTENSIONS:
        raise Exception("invalid image path: {}".format(X_img_path))

    if knn_clf is None and model_save_path == "":
        raise Exception("must supply knn classifier either thourgh knn_clf or model_save_path")

    if knn_clf is None:
        with open(model_save_path, 'rb') as f:
            knn_clf = pickle.load(f)

    X_img = face_recognition.load_image_file(X_img_path)
    X_faces_loc = face_recognition.face_locations(X_img)
    if len(X_faces_loc) == 0:
        return []

    faces_encodings = face_recognition.face_encodings(X_img, known_face_locations=X_faces_loc,num_jitters=1)


    closest_distances = knn_clf.kneighbors(faces_encodings, n_neighbors=1)

    is_recognized = [closest_distances[0][i][0] <= DIST_THRESH for i in range(len(X_faces_loc))]

    # predict classes and cull classifications that are not with high confidence
    return [(pred) if rec else ("Unknown") for pred, loc, rec in zip(knn_clf.predict(faces_encodings), X_faces_loc, is_recognized)]

if __name__ == "__main__":
    browseDir = askdirectory(title = "Choose the Folder")
    date = browseDir[-8:]
    #print(date)
    path = browseDir
    path = str(path)
    book = load_workbook(filename = 'Attendance Records/Attendance.xlsx')
    #book = Workbook()
    sheet = book.active

    sheet.cell(row = 1, column = (len(sheet['1'])+1)).value = date
   
    for row in range(2,(len(sheet['A'])+1)):
        sheet.cell(row = row, column = len(sheet['1'])).value = "A"    

    max_col = len(sheet['1'])
    for img_path in listdir(path):
        preds = predict(join(path, img_path) ,model_save_path=("model/knn_model.pkl"))
        print(preds)
        preds = str(preds)
        if preds != "[]" and preds != "['Unknown']" :
            roll_no = preds[2:6]
            roll_no = int(roll_no)
            for row in range(2,(len(sheet['A'])+1)):
                if sheet.cell(row = row, column = 1).value == roll_no:
                    sheet.cell(row = row, column = max_col).value = "P"
    
    tkinter.messagebox.showinfo('SUCCESS!',"Attendance Marked Successfully")
    book.save("Attendance Records/Attendance.xlsx")

